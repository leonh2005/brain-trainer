from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "黃金版配置"

# ── Colors ──
BG_DARK   = "0D1117"
GOLD      = "C8A96E"
GOLD_SOFT = "F5EDD8"
GREEN_BG  = "D6F4E5"
GREEN_TXT = "1A7A45"
SURFACE   = "F8F9FA"
HEADER_BG = "1C1F26"
BORDER_C  = "D0D7DE"
BLUE_TXT  = "0000FF"
BLACK_TXT = "000000"

def side(color=BORDER_C):
    return Side(style='thin', color=color)

def border(color=BORDER_C):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value, bold=False, color=BLACK_TXT, bg=None, align='left', size=11, wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = border()
    return c

# ═══════════════════════════════
# SECTION 1: 標題
# ═══════════════════════════════
ws.merge_cells('A1:H1')
c = ws['A1']
c.value = "黃金版投資組合配置"
c.font = Font(name='Arial', bold=True, size=16, color="FFFFFF")
c.fill = PatternFill('solid', start_color=HEADER_BG)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:H2')
c = ws['A2']
c.value = "穩健成長型 · 大漲+24.0% / 大跌-15.9% / 賺賠比 1.51"
c.font = Font(name='Arial', size=10, color=GOLD, italic=True)
c.fill = PatternFill('solid', start_color=HEADER_BG)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# ═══════════════════════════════
# SECTION 2: 配置主表
# ═══════════════════════════════
ws.row_dimensions[3].height = 8  # spacer

HEADERS = ['標的', '名稱', '配置（萬）', '佔比', '五年報酬', '如何買入', '交易所', '備註']
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill('solid', start_color="2D3748")
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border("2D3748")
ws.row_dimensions[4].height = 22

HOLDINGS = [
    ('006208', '富邦台50',        140, '=C5/280',  '+149.1%', '台股券商直接下單',           'TWSE',  '台積電佔比高，台股核心'),
    ('00881',  '國泰永續高股息',  140, '=C6/280',  '+189.1%', '台股券商直接下單',           'TWSE',  '2026初高配息，填息中'),
    ('VWRA',   '先鋒全球股票',    140, '=C7/280',  '+50.0%',  'IBKR / 複委託',              'LSE',   '全球分散，累積型不配息'),
    ('GRID',   '全球電力基建',    105, '=C8/280',  '+93.2%',  'IBKR / 複委託',              'NYSE',  'AI基建結構性需求'),
    ('00635U', '元大黃金',         70, '=C9/280',  '+93.3%',  '台股券商直接下單',           'TWSE',  '通膨避險，非相關資產'),
    ('XLU',    '美國公用事業',     35, '=C10/280', '+65.6%',  'IBKR / 複委託',              'NYSE',  '防禦性，抗跌穩定器'),
    ('00864B', '中信美債0-1年',    35, '=C11/280', '+33.2%',  '台股券商直接下單',           'TWSE',  '極低波動，質押抵押品'),
]

for i, row_data in enumerate(HOLDINGS):
    r = 5 + i
    ws.row_dimensions[r].height = 20
    bg = "FFFFFF" if i % 2 == 0 else SURFACE

    ticker, name, amt, pct, ret5, how, exch, note = row_data

    cell(ws, r, 1, ticker,  bold=True, color="1A56DB", bg=bg, align='center')
    cell(ws, r, 2, name,    bg=bg)
    # amount - blue (hardcoded input)
    c = ws.cell(row=r, column=3, value=amt)
    c.font = Font(name='Arial', bold=False, color=BLUE_TXT, size=11)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    c.number_format = '#,##0'
    # pct - formula (black)
    c = ws.cell(row=r, column=4, value=pct)
    c.font = Font(name='Arial', color=BLACK_TXT, size=11)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    c.number_format = '0.0%'

    ret_color = GREEN_TXT if ret5.startswith('+') else "C0392B"
    cell(ws, r, 5, ret5,  color=ret_color, bg=bg, align='center', bold=True)
    cell(ws, r, 6, how,   bg=bg)
    cell(ws, r, 7, exch,  bg=bg, align='center')
    cell(ws, r, 8, note,  bg=bg, color="5A6078")

# 合計列
r = 12
ws.row_dimensions[r].height = 22
for col in range(1, 9):
    c = ws.cell(row=r, column=col)
    c.fill = PatternFill('solid', start_color="2D3748")
    c.border = border("2D3748")
    c.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')

ws['A12'] = '合計'
ws['C12'] = '=SUM(C5:C11)'
ws['C12'].number_format = '#,##0'
ws['C12'].font = Font(name='Arial', bold=True, color=GOLD, size=11)
ws['D12'] = '=SUM(D5:D11)'
ws['D12'].number_format = '0.0%'
ws['D12'].font = Font(name='Arial', bold=True, color=GOLD, size=11)

# ═══════════════════════════════
# SECTION 3: 分批投入計畫
# ═══════════════════════════════
ws.row_dimensions[13].height = 12  # spacer

ws.merge_cells('A14:H14')
c = ws['A14']
c.value = "分批投入計畫"
c.font = Font(name='Arial', bold=True, size=12, color="FFFFFF")
c.fill = PatternFill('solid', start_color=HEADER_BG)
c.alignment = Alignment(horizontal='left', vertical='center')
c.border = border(HEADER_BG)
ws.row_dimensions[14].height = 26

BATCH_HDR = ['批次', '預計時間', '金額（萬）', '佔總額', '優先標的', '買入邏輯', '', '']
for col, h in enumerate(BATCH_HDR[:6], 1):
    c = ws.cell(row=15, column=col, value=h)
    c.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill('solid', start_color="4A5568")
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border("4A5568")
ws.row_dimensions[15].height = 20

BATCHES = [
    ('第一批', '現在（2026/04）',  280, '=C16/700', '006208、00881、00635U、00864B', '台股四檔先建倉，熟悉且可立即執行'),
    ('第二批', '一個月後（05月）', 210, '=C17/700', 'VWRA、GRID',                    '觀察關稅情勢後，海外部位分批進場'),
    ('第三批', '兩個月後（06月）', 210, '=C18/700', 'XLU，補足各標的',               '完成全部建倉，依市況微調比例'),
]

for i, row_data in enumerate(BATCHES):
    r = 16 + i
    ws.row_dimensions[r].height = 22
    bg = GOLD_SOFT if i == 0 else ("FFFFFF" if i % 2 == 0 else SURFACE)
    batch, timing, amt, pct, targets, logic = row_data

    cell(ws, r, 1, batch,   bold=True, bg=bg, align='center', color=HEADER_BG)
    cell(ws, r, 2, timing,  bg=bg, align='center')
    c = ws.cell(row=r, column=3, value=amt)
    c.font = Font(name='Arial', color=BLUE_TXT, size=11)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    c.number_format = '#,##0'
    c = ws.cell(row=r, column=4, value=pct)
    c.font = Font(name='Arial', color=BLACK_TXT, size=11)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    c.number_format = '0.0%'
    cell(ws, r, 5, targets, bg=bg, wrap=True)
    cell(ws, r, 6, logic,   bg=bg, color="5A6078", wrap=True)

# 合計
r = 19
ws.row_dimensions[r].height = 20
for col in range(1, 7):
    c = ws.cell(row=r, column=col)
    c.fill = PatternFill('solid', start_color="4A5568")
    c.border = border("4A5568")
    c.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')
ws['A19'] = '合計'
ws['C19'] = '=SUM(C16:C18)'
ws['C19'].number_format = '#,##0'
ws['C19'].font = Font(name='Arial', bold=True, color=GOLD, size=11)
ws['D19'] = '=SUM(D16:D18)'
ws['D19'].number_format = '0.0%'
ws['D19'].font = Font(name='Arial', bold=True, color=GOLD, size=11)

# ═══════════════════════════════
# SECTION 4: 情境模擬
# ═══════════════════════════════
ws.row_dimensions[20].height = 12

ws.merge_cells('A21:H21')
c = ws['A21']
c.value = "情境模擬"
c.font = Font(name='Arial', bold=True, size=12, color="FFFFFF")
c.fill = PatternFill('solid', start_color=HEADER_BG)
c.alignment = Alignment(horizontal='left', vertical='center')
c.border = border(HEADER_BG)
ws.row_dimensions[21].height = 26

SIM_HDR = ['情境', '006208報酬', '組合預估報酬', '獲利/損失（萬）', '賺賠比', '', '', '']
for col, h in enumerate(SIM_HDR[:5], 1):
    c = ws.cell(row=22, column=col, value=h)
    c.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill('solid', start_color="4A5568")
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border("4A5568")
ws.row_dimensions[22].height = 20

SIMS = [
    ('大漲情境', '+35%', '+24.0%', '=700*0.24', GREEN_BG, GREEN_TXT),
    ('大跌情境', '-20%', '-15.9%', '=-700*0.159', "FDE8E8", "C0392B"),
]
for i, (label, ret206, port_ret, pnl, bg, tc) in enumerate(SIMS):
    r = 23 + i
    ws.row_dimensions[r].height = 22
    cell(ws, r, 1, label,    bold=True, bg=bg, color=tc, align='center')
    cell(ws, r, 2, ret206,   bg=bg, color=tc, align='center', bold=True)
    cell(ws, r, 3, port_ret, bg=bg, color=tc, align='center', bold=True)
    c = ws.cell(row=r, column=4, value=pnl)
    c.font = Font(name='Arial', bold=True, color=tc, size=11)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    c.number_format = '#,##0'

r = 25
ws.row_dimensions[r].height = 22
for col in range(1, 6):
    c = ws.cell(row=r, column=col)
    c.fill = PatternFill('solid', start_color="4A5568")
    c.border = border("4A5568")
    c.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')
ws['A25'] = '賺賠比'
ws['E25'] = 1.51
ws['E25'].font = Font(name='Arial', bold=True, color=GOLD, size=13)
ws['E25'].alignment = Alignment(horizontal='center', vertical='center')
ws['E25'].number_format = '0.00'

# ═══════════════════════════════
# Column widths
# ═══════════════════════════════
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 24

# Freeze pane
ws.freeze_panes = 'A5'

out = '/Users/steven/CCProject/portfolio_黃金版.xlsx'
wb.save(out)
print(f"✓ 已儲存：{out}")
